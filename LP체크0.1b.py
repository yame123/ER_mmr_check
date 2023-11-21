import sys
from tabnanny import check
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QCheckBox, QPushButton, QHBoxLayout, QVBoxLayout
from PyQt5.QtCore import Qt
import openpyxl
from openpyxl import Workbook
import json
import requests
import time

class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()
        self.excel_address=''
        self.api_key=''
        self.standard_date=0
        self.played_game=0
        self.seasonid=0
        self.teammode=[0,0,0]
        self.done_state=''

    def initUI(self):
        self.lbl = QLabel(self)

        self.qle0_sheetaddress = QLineEdit(self)
        qlabel0=QLabel('엑셀 시트 주소',self)
        self.qle0_sheetaddress.textChanged[str].connect(self.qle0changed)

        haddressbox=QHBoxLayout()
        haddressbox.addWidget(qlabel0)
        haddressbox.addWidget(self.qle0_sheetaddress)
        haddressbox.addStretch(1)
        
        self.qle1_apikey = QLineEdit(self)
        qlabel1=QLabel('API key',self)
        self.qle1_apikey.textChanged[str].connect(self.qle1changed)

        hapikeybox=QHBoxLayout()
        hapikeybox.addWidget(qlabel1)
        hapikeybox.addWidget(self.qle1_apikey)
        hapikeybox.addStretch(1)

        self.qle3_seasonid = QLineEdit(self)
        
        self.qle3_seasonid.textChanged[str].connect(self.qle3changed)
        qlabel3=QLabel('season Id',self)
        hseasonidbox=QHBoxLayout()
        hseasonidbox.addWidget(qlabel3)
        hseasonidbox.addWidget(self.qle3_seasonid)
        hseasonidbox.addStretch(1)

        playcheck = QCheckBox('play check',self)
        playcheck.stateChanged.connect(self.playedgamechecking)
        
        self.qle2_standardate = QLineEdit(self)
        self.qle2_standardate.setMaxLength(10)
        self.qle2_standardate.textChanged[str].connect(self.qle2changed)
        qlabel2=QLabel('Standard Date',self)
        hstandarddatebox=QHBoxLayout()
        hstandarddatebox.addWidget(qlabel2)
        hstandarddatebox.addWidget(self.qle2_standardate)
        hstandarddatebox.addStretch(1)

        solocheck = QCheckBox('Solo',self)
        solocheck.stateChanged.connect(self.solochecking)
        duocheck = QCheckBox('Duo',self)
        duocheck.stateChanged.connect(self.duochecking)
        squadcheck = QCheckBox('Squad',self)
        squadcheck.stateChanged.connect(self.squadchecking)
        button = QPushButton('run')
        button.clicked.connect(self.run)
        
        hcheckboxs=QHBoxLayout()
        hcheckboxs.addWidget(solocheck)
        hcheckboxs.addWidget(duocheck)
        hcheckboxs.addWidget(squadcheck)
        hcheckboxs.addStretch(1)

        allboxs=QVBoxLayout()
        allboxs.addStretch(1)
        allboxs.addLayout(haddressbox)
        allboxs.addLayout(hapikeybox)
        allboxs.addLayout(hseasonidbox)
        allboxs.addWidget(playcheck)
        allboxs.addLayout(hstandarddatebox)
        allboxs.addLayout(hcheckboxs)
        allboxs.addWidget(self.lbl)
        allboxs.addWidget(button)
        allboxs.addStretch(1)

        self.setLayout(allboxs)

        self.setWindowTitle('Auto LP')
        self.setGeometry(300, 300, 300, 200)
        self.show()
    def qle0changed(self, text):
        self.excel_address=text
    def qle1changed(self, text):
        self.api_key=text
    def qle2changed(self, text):
        try:
            self.standard_date=int(text)
        except:
            if text=='':
                self.standard_date=0
            self.qle2_standardate.setText(text[:-1])
    def qle3changed(self, text):
        try:
            self.seasonid=int(text)
        except:
            if text=='':
                self.seasonid=0
            self.qle3_seasonid.setText(text[:-1])
        
    def solochecking(self, state):
        if state == Qt.Checked:
            self.teammode[0]=1
        else:
            self.teammode[0]=0
    def duochecking(self, state):
        if state == Qt.Checked:
            self.teammode[1]=1
        else:
            self.teammode[1]=0
    def squadchecking(self, state):
        if state == Qt.Checked:
            self.teammode[2]=1
        else:
            self.teammode[2]=0
    def playedgamechecking(self, state):
        if state == Qt.Checked:
            self.played_game=1
        else:
            self.played_game=0
    
    
    def run(self):
        self.lbl.setText('Running...')
        explace = self.excel_address.replace('\\','/').replace('\u202a','')
        size = 0
        exc = openpyxl.load_workbook(explace)
        sheet = exc['MMR']
        nickname_group = []
        usernum_group=[]
        while(1):
            if(sheet[f'A{size+3}'].value == None):
                break
            nickname_group.append(sheet[f'A{size+3}'].value)
            size=size+1
        print("size ",size)
        endpoint = 'https://open-api.bser.io'
        API_Key = self.api_key
        nickname = '/v1/user/nickname/'
        game='/v1/user/games/'
        headers = {'accept' : 'application/json','x-api-key':API_Key}

        standard = self.standard_date
        playedgamecheck = self.played_game
        season = self.seasonid
        team = self.teammode

        for j in range(size):
            params_nickname={('query',nickname_group[j]),}
            while(1):
                r = requests.get((endpoint+nickname),headers = headers,params = params_nickname)
                data = r.json()
                
                if data['message'] == 'Success': #Success
                    usernum_test = int(data["user"]['userNum'])
                    break
                elif data['message'] == 'Not Found': #Not Found
                    usernum_test = -0.5
                    break
                else:
                    pass
            usernum_group.append(usernum_test)
            print('name',j+1)

        for matchingTeamMode in range(3):
            if team[matchingTeamMode]==1:
                for i in range(size):
                    userNum = usernum_group[i]
                    rank = f'/v1/rank/{userNum}/{season}/{matchingTeamMode+1}'
                    while(1):
                        r = requests.get(endpoint+rank,headers = headers,)
                        data = r.json()
                        if data['message'] == 'Success':
                            mmr = int(data['userRank']['mmr'])
                            break
                        elif data['message'] == 'Not Found': #Not Found
                            mmr = '없는 계정'
                            break
                        else:
                            pass
                    sheet.cell(i+3,matchingTeamMode+2).value=mmr
                    print(matchingTeamMode+1,'mmr',i+1)

        if playedgamecheck == 1:   
            for i in range(size):
                while(1):
                    r = requests.get(endpoint+game+str(usernum_group[i]),headers = headers,)
                    data = r.json()
                    if data['message'] == 'Success' :
                        tmp=data["userGames"][0]['startDtm']
                        playtime = int(tmp[:4]+tmp[5:7]+tmp[8:10]+tmp[11:13])
                        if playtime <standard:
                            mmr = ''
                        else:
                            mmr = '기준 시각 이후 플레이 기록 존재'
                        break
                    elif usernum_group[i]<0:
                        mmr = '없는 계정'
                        break
                    else:
                        pass
                sheet.cell(i+3,matchingTeamMode+5).value=mmr
                print(matchingTeamMode+1,'한판 확인',i+1)
                
        exc.save(explace)
        print("Check the file")
        self.lbl.setText('Done!')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())